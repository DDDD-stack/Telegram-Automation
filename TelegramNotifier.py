import re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook

senderEmail = "shqipg46@gmail.com"
appPassword = "drlb opqx dayg vovv"
receiver = "drobi840@gmail.com"
excel = "csroi.xlsx"


def parseRoi(val):
    match = re.search(r"-?[\d]+\.?\d*", str(val))
    return float(match.group()) if match else -float("inf")


def loadTop3():
    wb = load_workbook(excel)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    roiCol   = next((i+1 for i, h in enumerate(headers) if h and "Investing ROI" in str(h)), None)
    priceCol = next((i+1 for i, h in enumerate(headers) if h and "Price" in str(h)), None)
    linkCol  = next((i+1 for i, h in enumerate(headers) if h and "Link" in str(h)), None)

    items = []
    for row in range(2, ws.max_row + 1):
        roi  = ws.cell(row=row, column=roiCol).value
        info = ws.cell(row=row, column=priceCol).value
        link = str(ws.cell(row=row, column=linkCol).value)
        if not roi or not info or str(info).strip() in ("", "N/A"):
            continue
        if not re.search(r"\d", str(roi)):
            continue
        items.append({"roi": parseRoi(roi), "info": str(info).strip(), "link": link})

    return sorted(items, key=lambda x: x["roi"], reverse=True)[:3]


def main():
    top3 = loadTop3()

    plain = "TOP 3 CS2 ITEMS BY INVESTING ROI\n\n"
    html  = "TOP 3 CS2 ITEMS BY INVESTING ROI<br><br>"
    for i, item in enumerate(top3, 1):
        plain += f"{i}. {item['info']}\n   Link: {item['link']}\n\n"
        html  += f"<b>{i}. {item['info']}</b><br><a href='{item['link']}'>View Item</a><br><br>"

    print(plain)

    msg = MIMEMultipart("alternative")
    msg["Subject"] = "Top 3 CS2 Investments"
    msg["From"]    = senderEmail
    msg["To"]      = receiver
    msg.attach(MIMEText(plain, "plain"))
    msg.attach(MIMEText(html,  "html"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(senderEmail, appPassword)
        server.send_message(msg)
        print("Email sent!")

if __name__ == "__main__":
    main()