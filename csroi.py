from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

excel = "csroi.xlsx"
tabs = [("Cases", "Case"), ("Stickers", "Sticker"), ("Armory", "Armory")]


def acceptCookies(page):
    try:
        page.click("button:has-text('Agree')", timeout=4000)
        page.wait_for_timeout(800)
    except Exception:
        pass


def switchTab(page, value):
    btn = page.query_selector(f"button[value='{value}']")
    # Only click if not already selected
    if btn and btn.get_attribute("aria-pressed") != "true":
        btn.click()
        page.wait_for_timeout(2500)
    else:
        print(f"  Tab '{value}' already active, skipping click.")


def selectInvestRoi(page):
    try:
        page.click("div[class*='MuiSelect']:has-text('INVEST ROI'), div[class*='MuiSelect']:has-text('UNBOX ROI')", timeout=3000)
        page.wait_for_timeout(800)
        page.click("li[data-value='InvestROI'], li:has-text('INVEST ROI')", timeout=2000)
        page.wait_for_timeout(1500)
    except Exception:
        pass


def scrapeTop10(page):
    page.wait_for_timeout(1000)
    cards = page.query_selector_all("div.MuiPaper-root.MuiCard-root")[:10]
    results = []
    for card in cards:
        lines = [l.strip() for l in card.inner_text().split("\n") if l.strip()]
        name, price, roi, profit = "N/A", "N/A", "N/A", "N/A"
        for i, line in enumerate(lines):
            if ("€" in line or "$" in line) and price == "N/A":
                price = line
            if line == "Investing ROI" and i + 1 < len(lines):
                roi = lines[i + 1]
            if line == "Profit" and i + 1 < len(lines):
                profit = lines[i + 1]
        if lines:
            name = lines[0]
        results.append({"Name": name, "Price": price, "Investing ROI": roi, "Profit Chance": profit})
    return results


def writeExcel(allData):
    wb = Workbook()
    ws = wb.active
    ws.title = "CSROI Top 10"
    sectionColors = {"Cases": "C00000", "Stickers": "7030A0", "Armory": "375623"}

    for col, (header, width) in enumerate(zip(
        ["#", "Name", "Price", "Investing ROI (1M)", "Profit Chance"],
        [4, 42, 12, 18, 16]
    ), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    row = 2
    for section, items in allData.items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=f"  {section.upper()}")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=sectionColors[section])
        row += 1
        for i, item in enumerate(items):
            for col, val in enumerate([i+1, item["Name"], item["Price"], item["Investing ROI"], item["Profit Chance"]], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.fill = PatternFill("solid", fgColor="DCE6F1" if i % 2 == 0 else "FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            row += 1
        row += 1

    wb.save(excel)
    print(f"Saved to {excel}")


def main():
    allData = {}
    with sync_playwright() as p:
        browser = p.webkit.launch(headless=False)
        page = browser.new_page()
        page.goto("https://csroi.com", wait_until="networkidle")
        page.wait_for_timeout(3000)
        acceptCookies(page)

        for label, value in tabs:
            print(f"Scraping {label}...")
            switchTab(page, value)
            selectInvestRoi(page)
            allData[label] = scrapeTop10(page)
            print(f"  Got {len(allData[label])} items.")

        browser.close()

    writeExcel(allData)


if __name__ == "__main__":
    main()